import random
from openpyxl import Workbook

head_sector = 0
average_search_time = 6.46
rotational_delay = 4.17
transfer_delay = 0.13
start_end_delay = 1.0
disk_sector = 65535
sector_move_count = 4000
isRandom = True 
show_each_request_detail = True  

class Request:
    def __init__(self, sector, time):
        self.sector = sector
        self.time = time

def bubble_sort(arr):
    n = len(arr)
    for i in range(n - 1):
        for j in range(n - i - 1):
            if arr[j].time > arr[j + 1].time:
                arr[j], arr[j + 1] = arr[j + 1], arr[j]

def make_random_request():
    global head_sector
    head_sector = random.randint(0, disk_sector // sector_move_count) * sector_move_count

    number = random.randint(2, 10)  # تعداد ریکوئست ها بین اعداد 2 تا 10
    requests = []

    for _ in range(number):
        sector = random.randint(0, disk_sector // sector_move_count) * sector_move_count
        time = random.randint(0, 9) * 10
        requests.append(Request(sector, time))

    bubble_sort(requests)
    return requests


def calculate_eval(request_sector, current_sector, total_time):
    total_time += rotational_delay + transfer_delay
    if current_sector != request_sector:
        total_time += (abs(current_sector - request_sector) / sector_move_count)+ start_end_delay

    if show_each_request_detail:
        print(request_sector, total_time)

    return total_time


def request_go_down_eval(requests, current_sector):
    max_sector = -1
    index = -1

    for i, request in enumerate(requests):
        if current_sector >= request.sector > max_sector:
            max_sector = request.sector
            index = i

    return index


def request_go_up_eval(requests, current_sector):
    min_sector = disk_sector
    index = -1

    for i, request in enumerate(requests):
        if current_sector <= request.sector < min_sector:
            min_sector = request.sector
            index = i

    return index


def get_requests_within_time(requests, time):
    return [req for req in requests if req.time <= time]


def elevator(requests):
    if not isRandom:
        print("Elevator Results:")
    available_requests = list(requests)
    total_time = 0.0
    current_sector = head_sector
    direction_is_up = True

    while available_requests:
        requests_in_time = get_requests_within_time(available_requests, total_time)

        if not requests_in_time:
            total_time = available_requests[0].time
            continue

        if direction_is_up:
            index = request_go_up_eval(requests_in_time, current_sector)
            if index == -1:
                index = request_go_down_eval(requests_in_time, current_sector)
                direction_is_up = False
        else:
            index = request_go_down_eval(requests_in_time, current_sector)
            if index == -1:
                index = request_go_up_eval(requests_in_time, current_sector)
                direction_is_up = True

        total_time = calculate_eval(requests_in_time[index].sector, current_sector, total_time)
        current_sector = requests_in_time[index].sector
        available_requests.pop(index)

    if isRandom:
        return total_time
    else:
        print("Total time is:", total_time)
        print()


def fcfs(requests):
    if not isRandom:
        print("FCFS Results:")
    current_sector = head_sector
    total_time = 0.0

    for request in requests:
        if request.time > total_time:
            total_time = request.time

        total_time += rotational_delay + transfer_delay
        if current_sector != request.sector:
            total_time += (abs(current_sector - request.sector) / sector_move_count) + 1

        current_sector = request.sector
        if show_each_request_detail:
            print(request.sector, total_time)
    if isRandom:
        return total_time
    else:
        print("Total time is:", total_time)
        print()


def get_requests():
    global head_sector
    head_sector = int(input("enter starting head sector: "))
    number = int(input("enter number of requests: "))
    requests = []

    print("Enter requests in order:")
    for _ in range(number):
        line = input().split()
        requests.append(Request(int(line[0]), int(line[1])))

    return requests


def choose():
    print("welcome :)")
    print("1. Manual quantity")
    print("2. Random")
    get = int(input("-> "))
    return get == 1


def main():
    global isRandom, show_each_request_detail

    if choose():
        isRandom = False
        requests = get_requests()
        fcfs(requests)
        elevator(requests)
    else:
        show_each_request_detail = False
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "FCFS Time"
        ws.cell(row=1, column=2).value = "Elevator Time"

        for i in range(1000):
            requests = make_random_request()
            total_fcfs_time = fcfs(requests)
            total_elevator_time = elevator(requests)
            ws.cell(row=i + 2, column=1).value = total_fcfs_time
            ws.cell(row=i + 2, column=2).value = total_elevator_time
        wb.save("plot.xlsx")


if __name__ == "__main__":
    main()
